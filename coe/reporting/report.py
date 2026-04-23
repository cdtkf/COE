#!/usr/bin/env python3
"""
report_generator.py — Generate ranked opportunity reports from scored data.

Produces CSV and/or XLSX files with scored opportunities, color-coded by
match quality, with filtering options.

Usage:
    python report_generator.py                         # Generate both CSV and XLSX
    python report_generator.py --format csv            # CSV only
    python report_generator.py --min-score 50          # Only scores >= 50
    python report_generator.py --days 7                # Only last 7 days
    python report_generator.py --active                # Only active/open opportunities
    python report_generator.py --office 36C10B         # Filter by office
    python report_generator.py --output my_report      # Custom output filename
"""

import argparse
import csv
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import yaml

from coe.puller.sqlite_db import Database

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.yaml"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


def load_config() -> dict:
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


def generate_csv(rows: list, output_path: Path):
    """Generate a CSV report from scored opportunity rows."""
    headers = [
        "Rank", "Overall Score", "Title", "Solicitation Number",
        "Office", "Office Code", "Agency", "Notice Type", "NAICS Code",
        "Set-Aside", "Response Deadline", "Posted Date",
        "Domain Score", "Capability Score", "NAICS Score", "Set-Aside Score",
        "Rationale", "Best Matching Proposals", "Key Alignment Factors",
        "Risk Factors", "SAM.gov Link", "Active",
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for rank, row in enumerate(rows, 1):
            # Parse JSON fields
            matched = _parse_json_field(row["matched_profiles"])
            alignment = _parse_json_field(row["key_alignment_factors"])
            risks = _parse_json_field(row["risk_factors"])

            writer.writerow([
                rank,
                row["overall_score"],
                row["title"],
                row["solicitation_number"] or "",
                row["office"] or "",
                row["office_code"] or "",
                row["department"] or "",
                row["base_type"] or row["notice_type"] or "",
                row["naics_code"] or "",
                row["set_aside_type"] or "Unrestricted",
                row["response_deadline"] or "",
                row["posted_date"] or "",
                row["domain_score"],
                row["capability_score"],
                row["naics_score"],
                row["set_aside_fit"],
                row["rationale"] or "",
                "; ".join(matched) if isinstance(matched, list) else str(matched),
                "; ".join(alignment) if isinstance(alignment, list) else str(alignment),
                "; ".join(risks) if isinstance(risks, list) else str(risks),
                row["description_url"] or "",
                row["active"] or "",
            ])

    log.info(f"CSV report saved: {output_path}")


def generate_xlsx(rows: list, output_path: Path):
    """Generate an XLSX report with conditional formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunity Matches"

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Score color fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_font = Font(name="Arial", color="006100", bold=True)
    yellow_font = Font(name="Arial", color="9C6500", bold=True)
    red_font = Font(name="Arial", color="9C0006")

    # Headers
    headers = [
        "Rank", "Score", "Title", "Solicitation #",
        "Office", "Agency", "Notice Type", "NAICS",
        "Set-Aside", "Deadline", "Posted",
        "Domain", "Capability", "NAICS Fit", "Set-Aside Fit",
        "Rationale", "Best Matching Proposals", "SAM.gov Link",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for rank, row in enumerate(rows, 1):
        matched = _parse_json_field(row["matched_profiles"])
        data_row = rank + 1

        values = [
            rank,
            row["overall_score"],
            row["title"],
            row["solicitation_number"] or "",
            f"{row['office'] or ''} ({row['office_code'] or ''})",
            row["department"] or "",
            row["base_type"] or row["notice_type"] or "",
            row["naics_code"] or "",
            row["set_aside_type"] or "Unrestricted",
            row["response_deadline"] or "",
            row["posted_date"] or "",
            row["domain_score"],
            row["capability_score"],
            row["naics_score"],
            row["set_aside_fit"],
            row["rationale"] or "",
            "; ".join(matched) if isinstance(matched, list) else str(matched),
            row["description_url"] or "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in [3, 16, 17]))

        # Color-code the score cell
        score_cell = ws.cell(row=data_row, column=2)
        score = row["overall_score"]
        if score >= 70:
            score_cell.fill = green_fill
            score_cell.font = green_font
        elif score >= 40:
            score_cell.fill = yellow_fill
            score_cell.font = yellow_font
        else:
            score_cell.fill = red_fill
            score_cell.font = red_font

        # Also color sub-score cells (columns 12-15)
        for sub_col in [12, 13, 14, 15]:
            sub_cell = ws.cell(row=data_row, column=sub_col)
            sub_val = sub_cell.value or 0
            if sub_val >= 70:
                sub_cell.fill = green_fill
            elif sub_val >= 40:
                sub_cell.fill = yellow_fill
            else:
                sub_cell.fill = red_fill

        # Make SAM.gov link clickable
        link_cell = ws.cell(row=data_row, column=18)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.font = Font(name="Arial", color="0563C1", underline="single")

    # Column widths
    col_widths = {
        1: 6, 2: 8, 3: 45, 4: 20, 5: 30, 6: 30, 7: 15, 8: 10,
        9: 18, 10: 14, 11: 14, 12: 9, 13: 11, 14: 10, 15: 12,
        16: 50, 17: 30, 18: 35,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="SAM.gov Opportunity Match Report").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"Total Scored: {len(rows)}")

    high = sum(1 for r in rows if r["overall_score"] >= 70)
    med = sum(1 for r in rows if 40 <= r["overall_score"] < 70)
    low = sum(1 for r in rows if r["overall_score"] < 40)
    avg = sum(r["overall_score"] for r in rows) / len(rows) if rows else 0

    ws2.cell(row=4, column=1, value=f"High Match (70+): {high}")
    ws2.cell(row=4, column=1).font = Font(color="006100", bold=True)
    ws2.cell(row=5, column=1, value=f"Medium Match (40-69): {med}")
    ws2.cell(row=5, column=1).font = Font(color="9C6500", bold=True)
    ws2.cell(row=6, column=1, value=f"Low Match (<40): {low}")
    ws2.cell(row=6, column=1).font = Font(color="9C0006")
    ws2.cell(row=7, column=1, value=f"Average Score: {avg:.1f}")

    ws2.column_dimensions["A"].width = 40

    wb.save(output_path)
    log.info(f"XLSX report saved: {output_path}")


def _parse_json_field(value):
    """Parse a JSON string field, returning the parsed value or the original string."""
    if not value:
        return []
    if isinstance(value, list):
        return value
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return value


def main():
    parser = argparse.ArgumentParser(description="Generate ranked opportunity match reports")
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="both",
                        help="Output format (default: both)")
    parser.add_argument("--min-score", type=int, default=None,
                        help="Minimum score threshold (default: from config or 0)")
    parser.add_argument("--days", type=int, default=None,
                        help="Only include opportunities posted in last N days")
    parser.add_argument("--active", action="store_true",
                        help="Only include active/open opportunities")
    parser.add_argument("--office", default=None,
                        help="Filter by office code (e.g., 36C10B)")
    parser.add_argument("--output", default=None,
                        help="Output filename (without extension)")
    args = parser.parse_args()

    config = load_config()
    matching_cfg = config.get("matching", {})

    min_score = args.min_score if args.min_score is not None else matching_cfg.get("min_score_for_report", 0)

    db_path = config.get("settings", {}).get("database", "opportunities.db")
    if not Path(db_path).is_absolute():
        db_path = str(SCRIPT_DIR / db_path)

    db = Database(db_path)

    # Query scored opportunities
    rows = db.get_scored_opportunities(
        min_score=min_score,
        office_code=args.office,
        days=args.days,
        active_only=args.active,
    )

    if not rows:
        log.warning("No scored opportunities found matching your filters.")
        log.info("Run matcher.py first to score opportunities.")
        stats = db.get_scoring_stats()
        log.info(f"DB status: {stats}")
        db.close()
        return

    log.info(f"Found {len(rows)} scored opportunities (min score: {min_score})")

    # Generate output filename
    date_str = datetime.now().strftime("%Y%m%d")
    base_name = args.output or f"opportunity_matches_{date_str}"

    if args.format in ("csv", "both"):
        generate_csv(rows, SCRIPT_DIR / f"{base_name}.csv")

    if args.format in ("xlsx", "both"):
        generate_xlsx(rows, SCRIPT_DIR / f"{base_name}.xlsx")

    # Print top matches summary
    log.info(f"\nTop matches:")
    for i, row in enumerate(rows[:10], 1):
        score = row["overall_score"]
        emoji = "🟢" if score >= 70 else ("🟡" if score >= 40 else "🔴")
        log.info(f"  {emoji} {i}. [{score}] {(row['title'] or 'Untitled')[:65]}")

    db.close()


if __name__ == "__main__":
    main()
